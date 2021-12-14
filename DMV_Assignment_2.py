# import packages that we need
import pandas as pd
import numpy as np
from scipy import stats
import seaborn
import statsmodels.formula.api as stats_formula
import matplotlib.pyplot as pyplot
import openpyxl

# read from dataset
data = pd.read_csv('dataset/marscrater_pds.csv', low_memory=False)

# bug fix for display formats to avoid run time errors
pd.set_option('display.float_format', lambda x:'%f'%x)

print()
print('==================')
print('Data Information: ')
print('==================')
# number of observations (rows)
print('Number of records: ' + str(len(data)))
# number of variables (columns)
print('Number of varaibles: ' + str(len(data.columns)))

# checking the variable data type
print('Crater ID Variable Type: ' + str(data['DIAM_CIRCLE_IMAGE'].dtype))
print()

# setting variables to numeric
# data['DIAM_CIRCLE_IMAGE'] = pd.to_numeric(data['DIAM_CIRCLE_IMAGE'])
# print('Diameter Variable Type: ' + str(data['DIAM_CIRCLE_IMAGE'].dtype))

# counts and percentages (i.e. frequency distributions) for this variable
print()
print('Value counts for varaible - DAIM_CIRCLE_IMAGE: ')
count_1 = data['DIAM_CIRCLE_IMAGE'].value_counts(sort=False)
print(count_1)
print()

print()
print('Percentage for varaible - DAIM_CIRCLE_IMAGE: ')
percentage_1 = data['DIAM_CIRCLE_IMAGE'].value_counts(sort=False, normalize=True)
print(percentage_1)
print()

print()
print('Frequency  for varaible - DAIM_CIRCLE_IMAGE: ')
frequency_1 = data.groupby('DIAM_CIRCLE_IMAGE').size()
print(frequency_1)
print()

# subset data to craters with diameter between 50 and 80 kms
subset_1 = data[(data['DIAM_CIRCLE_IMAGE'] >= 50) & (data['DIAM_CIRCLE_IMAGE'] <= 80)]

subset_2 = subset_1.copy()

# frequency distributions on new subset_2 data frame
print()
print('Counts for Diameter Circle with subset: ')
count_2 = subset_2['DIAM_CIRCLE_IMAGE'].value_counts(sort=False)
print(count_2)
print()

# upper-case all DataFrame column names - place afer code for loading data aboave
data.columns = list(map(str.upper, data.columns))

# WORKING MORE
data['DIAM_CIRCLE_IMAGE'] = pd.to_numeric(data['DIAM_CIRCLE_IMAGE'], errors='coerce')
data['DEPTH_RIMFLOOR_TOPOG'] = pd.to_numeric(data['DEPTH_RIMFLOOR_TOPOG'], errors='coerce')

working_subset_1 = data[
    (data['NUMBER_LAYERS'] > 0) & (data['DIAM_CIRCLE_IMAGE'] > 0) & (data['DIAM_CIRCLE_IMAGE'] <= 100)
    & (data['DEPTH_RIMFLOOR_TOPOG'] > 0) & (data['DEPTH_RIMFLOOR_TOPOG'] <= 3)
]

working_subset_2 = working_subset_1.copy()

working_subset_2['DIAM_CIRCLE_IMAGE_CENTER'] = (
        working_subset_2['DIAM_CIRCLE_IMAGE'] - working_subset_2['DIAM_CIRCLE_IMAGE'].mean()
)

# explanatory variable
print()
print('=====================================')
print('Statistics for Explanatory Variable: ')
print('=====================================')
print(working_subset_2[['DIAM_CIRCLE_IMAGE_CENTER']].describe())
print()

# regresion model
print()
print('==========================================================')
print('Regresion Model between crater DIAMETER and crater DEPTH: ')
print('==========================================================')
regresion_1 = stats_formula.ols('DEPTH_RIMFLOOR_TOPOG ~ DIAM_CIRCLE_IMAGE_CENTER', data=working_subset_2).fit()
print(regresion_1.summary())
print()

slope, intercept, r_value, p_value, sdt_err = stats.linregress(working_subset_2['DIAM_CIRCLE_IMAGE_CENTER'], working_subset_2['DEPTH_RIMFLOOR_TOPOG'])

graph_1 = seaborn.lmplot(x='DIAM_CIRCLE_IMAGE_CENTER', y='DEPTH_RIMFLOOR_TOPOG', height=9, fit_reg=True, line_kws={'color':'red'}, data=working_subset_2)

pyplot.xlabel('CRATER DIAMETER CENTER (km)')
pyplot.ylabel('CRATER DEPTH (km)')
pyplot.title('Relationship between crater diameter and crater depth', fontsize=20, fontweight='bold')
graph_1.set(xlim=(-20, 80))
graph_1.set(ylim=(0, 4))
pyplot.xticks(np.arange(-20, 80+200, 20))
pyplot.yticks(np.arange(0, 4+1, 1))
pyplot.tick_params(axis='both', labelsize=19)
pyplot.show()
graph_1.savefig('output_regresion_model.png')


# ASSIGNMENT 2 for Course
data['NUMBER_LAYERS'] = pd.to_numeric(data['NUMBER_LAYERS'], errors='coerce')
data['DIAM_CIRCLE_IMAGE'] = pd.to_numeric(data['DIAM_CIRCLE_IMAGE'], errors='coerce')
data['DEPTH_RIMFLOOR_TOPOG'] = pd.to_numeric(data['DEPTH_RIMFLOOR_TOPOG'], errors='coerce')

# Frequency Distribution for NUMBER_LAWYERS Variable
print()
print('===========================================')
print('Value counts for varaible - NUMBER_LAYERS: ')
print('===========================================')
count_number_lawyers = data['NUMBER_LAYERS'].value_counts(sort=False)
print(count_number_lawyers)
print()

# Percentages for NUMBER_LAWYERS Variable
print()
print('==========================================')
print('Percentages for varaible - NUMBER_LAYERS: ')
print('==========================================')
percentages_number_lawyers = data['NUMBER_LAYERS'].value_counts(sort=False, normalize=True)
print(percentages_number_lawyers)
print()

# Frequency results
print()
print('=========')
print('Results: ')
print('=========')
data_frame_1 = pd.DataFrame(np.array(count_number_lawyers), index=count_number_lawyers.index, columns=['Frequency'])
data_frame_2 = pd.DataFrame(np.array(percentages_number_lawyers * 100), index=count_number_lawyers.index, columns=['Percentage'])
output_number_lawyers = pd.concat([data_frame_1, data_frame_2], axis=1)

output_number_lawyers['Frequency'] = output_number_lawyers.Frequency.cumsum()
output_number_lawyers['Percentage'] = output_number_lawyers.Percentage.cumsum()

output_number_lawyers.index.name = 'NUMBER_LAWYERS'
print(output_number_lawyers)
print()


# Frequency Distribution for MORPHOLOGY_EJECTA_3 Variable
# data['MORPHOLOGY_EJECTA_3'] = data['MORPHOLOGY_EJECTA_3'].apply(lambda x: np.nan if x == ' ' else x)
# data['MORPHOLOGY_EJECTA_3'] = pd.to_numeric(data['MORPHOLOGY_EJECTA_3'])

print()
print('=================================================')
print('Value counts for varaible - MORPHOLOGY_EJECTA_3: ')
print('=================================================')
count_morphology_ejecta_3 = data['MORPHOLOGY_EJECTA_3'].value_counts(sort=False, dropna=False)
print(count_morphology_ejecta_3)
print()

# Percentages for MORPHOLOGY_EJECTA_3 Variable
print()
print('================================================')
print('Percentages for varaible - MORPHOLOGY_EJECTA_3: ')
print('================================================')
percentages_morphology_ejecta_3 = data['MORPHOLOGY_EJECTA_3'].value_counts(sort=False, dropna=False, normalize=True)
print(percentages_morphology_ejecta_3)
print()

# Frequency results
print()
print('=========')
print('Results: ')
print('=========')
data_frame_3 = pd.DataFrame(
    np.array(count_morphology_ejecta_3),
    index=count_morphology_ejecta_3.index,
    columns=['Frequency']
)
data_frame_4 = pd.DataFrame(
    np.array(percentages_morphology_ejecta_3 * 100),
    index=percentages_morphology_ejecta_3.index,
    columns=['Percentage']
)
output_morphology_ejecta_3 = pd.concat([data_frame_3, data_frame_4], axis=1)

output_morphology_ejecta_3['Frequency'] = output_morphology_ejecta_3.Frequency.cumsum()
output_morphology_ejecta_3['Percentage'] = output_morphology_ejecta_3.Percentage.cumsum()

output_morphology_ejecta_3.index.name = 'MORPHOLOGY_EJECTA_3'
print(output_morphology_ejecta_3)
print()


# Frequency Distribution for MORPHOLOGY_EJECTA_2 Variable
# data['MORPHOLOGY_EJECTA_2'] = data['MORPHOLOGY_EJECTA_2'].apply(lambda x: np.nan if x == ' ' else x)
# data['MORPHOLOGY_EJECTA_2'] = pd.to_numeric(data['MORPHOLOGY_EJECTA_2'])

print()
print('=================================================')
print('Value counts for varaible - MORPHOLOGY_EJECTA_2: ')
print('=================================================')
count_morphology_ejecta_2 = data['MORPHOLOGY_EJECTA_2'].value_counts(sort=False, dropna=False)
print(count_morphology_ejecta_2)
print()

# Percentages for MORPHOLOGY_EJECTA_2 Variable
print()
print('================================================')
print('Percentages for varaible - MORPHOLOGY_EJECTA_2: ')
print('================================================')
percentages_morphology_ejecta_2 = data['MORPHOLOGY_EJECTA_2'].value_counts(sort=False, dropna=False, normalize=True)
print(percentages_morphology_ejecta_2)
print()

# Frequency results
print()
print('=========')
print('Results: ')
print('=========')
data_frame_5 = pd.DataFrame(
    np.array(count_morphology_ejecta_2),
    index=count_morphology_ejecta_2.index,
    columns=['Frequency']
)
data_frame_6 = pd.DataFrame(
    np.array(percentages_morphology_ejecta_2 * 100),
    index=percentages_morphology_ejecta_2.index,
    columns=['Percentage']
)
output_morphology_ejecta_2 = pd.concat([data_frame_5, data_frame_6], axis=1)

output_morphology_ejecta_2['Frequency'] = output_morphology_ejecta_2.Frequency.cumsum()
output_morphology_ejecta_2['Percentage'] = output_morphology_ejecta_2.Percentage.cumsum()

output_morphology_ejecta_2.index.name = 'MORPHOLOGY_EJECTA_2'
print(output_morphology_ejecta_2)
print()


# creating an workbook
excel_writer = pd.ExcelWriter('Mars_Craters_WorkBook_Assignment_2.xlsx')
output_number_lawyers.to_excel(
    excel_writer,
    sheet_name='NUMBER_LAWYERS',
    float_format='%0.4f',
    startrow=1,
    startcol=0
)
output_morphology_ejecta_3.to_excel(
    excel_writer,
    sheet_name='MORPHOLOGY_EJECTA_3',
    na_rep='NAN',
    float_format='%0.4f',
    startrow=1,
    startcol=0
)
output_morphology_ejecta_2.to_excel(
    excel_writer,
    sheet_name='MORPHOLOGY_EJECTA_2',
    na_rep='NAN',
    float_format='%0.4f',
    startrow=1,
    startcol=0
)
excel_writer.save()

wb = openpyxl.load_workbook('Mars_Craters_WorkBook_Assignment_2.xlsx')
sheets = wb.sheetnames

for i in range(len(sheets)):
    ws = wb.worksheets[i]
    variable_name = ws.cell(row=2, column=1).value
    title = 'Mars Crater Dataset Frequency Distribution for variable: ' + variable_name
    ws.cell(row=1, column=1).value = title.upper()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    wb.save('Mars_Craters_WorkBook_Assignment_2.xlsx')
