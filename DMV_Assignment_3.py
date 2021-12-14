# import libraries
import pandas
import numpy
import seaborn
import matplotlib.pyplot as pyplot
import matplotlib.ticker as ticker
import openpyxl

# read from the dataset
data = pandas.read_csv('dataset/marscrater_pds.csv', low_memory=False)

# bug fix for display formats to avoid run time errors
pandas.set_option('display.float_format', lambda x:'%f'%x)

# setting variables to numeric
data['LATITUDE_CIRCLE_IMAGE'] = pandas.to_numeric(data['LATITUDE_CIRCLE_IMAGE'])
data['LONGITUDE_CIRCLE_IMAGE'] = pandas.to_numeric(data['LONGITUDE_CIRCLE_IMAGE'])
data['DIAM_CIRCLE_IMAGE'] = pandas.to_numeric(data['DIAM_CIRCLE_IMAGE'])
data['DEPTH_RIMFLOOR_TOPOG'] = pandas.to_numeric(data['DEPTH_RIMFLOOR_TOPOG'])
data['NUMBER_LAYERS'] = pandas.to_numeric(data['NUMBER_LAYERS'])

# subset data with craters with diameter between 0 and 100 kms
subset_1 = data[(data['DIAM_CIRCLE_IMAGE'] > 0) & (data['DIAM_CIRCLE_IMAGE'] <= 100)]

# make a copy of subsetted data
subset_2 = subset_1.copy()

print('================================================')
print('Counts for original DIAM_CIRCLE_IMAGE Variable: ')
print('================================================')
count_1 = subset_2['DIAM_CIRCLE_IMAGE'].value_counts(sort=False, dropna=False)
print(count_1)
print()

# recode missing values to python missing (NaN)
subset_2['NUMBER_LAYERS'] = subset_2['NUMBER_LAYERS'].replace(4, numpy.nan)

# lets include a count of missing data:
print()
print('=========================================================================')
print('Counts for NUMBER_LAYERS = 3 set to NAN and number of missing requested: ')
print('=========================================================================')
count_2 = subset_2['NUMBER_LAYERS'].value_counts(sort=False, dropna=False)
print(count_2)
print()

# coding in valid data
# recode missing values to numeric value, in this example replace NaN with 10
subset_2['NUMBER_LAYERS'].fillna(10, inplace=True)
# recode 2 value as missing
subset_2['NUMBER_LAYERS'] = subset_2['NUMBER_LAYERS'].replace(2, numpy.nan)

print()
print('==========================================================')
print('NUMBER_LAYERS with Blanks recoded as 10 and 4 set to NAN: ')
print('==========================================================')
# check coding
check_code_1 = subset_2['NUMBER_LAYERS'].value_counts(sort=False, dropna=False)
print(check_code_1)
print()

print()
print('========================')
print('Describe NUMBER_LAYERS: ')
print('========================')
print()
describe_2 = subset_2["NUMBER_LAYERS"].describe()
print(describe_2)

# examining frequency distributions for NUMBER OF LAWYERS
print()
print('==========================')
print('Counts for NUMBER_LAYERS: ')
print('==========================')
count_5 = subset_2['NUMBER_LAYERS'].value_counts(sort=False)
print(count_5)
print()

print()
print('=============================')
print('Percentages for NUMBER_LAYERS')
print('=============================')
percentage_5 = subset_2['NUMBER_LAYERS'].value_counts(sort=False, normalize=True)
print(percentage_5)
print()

# quartile split (use qcut function & ask for 4 groups - gives you quartile split)
print()
print('=============================================')
print('DIAM_CIRCLE_IMAGE - 4 categories - quartiles:')
print('=============================================')
subset_2['DIAM_CIRCLE_IMAGE_GROUP_4'] = pandas.qcut(
    subset_2.DIAM_CIRCLE_IMAGE,
    4,
    labels=["1=0%tile", "2=25%tile", "3=50%tile", "4=75%tile"],
    duplicates='raise'
)
count_4 = subset_2['DIAM_CIRCLE_IMAGE_GROUP_4'].value_counts(sort=False, dropna=True)
print(count_4)
print()

# categorize quantitative variable based on customized splits using cut function
# splits into 3 groups (50-60, 60-70, 70-80)
print()
print('=============================')
print('DIAM_CIRCLE_IMAGE - 3 groups:')
print('=============================')
subset_2['DIAM_CIRCLE_IMAGE_GROUP_3'] = pandas.cut(subset_2.DIAM_CIRCLE_IMAGE, [0, 30, 60, 100])
count_5 = subset_2['DIAM_CIRCLE_IMAGE_GROUP_3'].value_counts(sort=False, dropna=True)
print(count_5)
print()

# crosstabs evaluating which DIAMETER were put into which DIAM_CIRCLE_IMAGE_GROUP_3
print()
print('==================================')
print('INFO DIAM_CIRCLE_IMAGE - 3 groups:')
print('==================================')
print(pandas.crosstab(subset_2['DIAM_CIRCLE_IMAGE_GROUP_3'], subset_2['DIAM_CIRCLE_IMAGE']))
print()

# frequency distribution for DIAM_CIRCLE_IMAGE_GROUP_3
print()
print('=====================================')
print('Counts for DIAM_CIRCLE_IMAGE_GROUP_3:')
print('=====================================')
count_10 = subset_2['DIAM_CIRCLE_IMAGE_GROUP_3'].value_counts(sort=False)
print(count_10)
print()
print()
print('==========================================')
print('Percentages for DIAM_CIRCLE_IMAGE_GROUP_3:')
print('==========================================')
percentage_10 = subset_2['DIAM_CIRCLE_IMAGE_GROUP_3'].value_counts(sort=False, normalize=True)
print(percentage_10)


# ASSIGNMENT 3 Continue
data['NUMBER_LAYERS'] = pandas.to_numeric(data['NUMBER_LAYERS'], errors='coerce')
data['DIAM_CIRCLE_IMAGE'] = pandas.to_numeric(data['DIAM_CIRCLE_IMAGE'], errors='coerce')
data['DEPTH_RIMFLOOR_TOPOG'] = pandas.to_numeric(data['DEPTH_RIMFLOOR_TOPOG'], errors='coerce')
data['LONGITUDE_CIRCLE_IMAGE'] = pandas.to_numeric(data['LONGITUDE_CIRCLE_IMAGE'], errors='coerce')
data['LATITUDE_CIRCLE_IMAGE'] = pandas.to_numeric(data['LATITUDE_CIRCLE_IMAGE'], errors='coerce')

graph_1 = seaborn.jointplot(
    x='LONGITUDE_CIRCLE_IMAGE', y='LATITUDE_CIRCLE_IMAGE', data=data, kind='reg', color='#774499', dropna=True,
    height=13, space=0.3, ratio=4, xlim=(-180, 180), ylim=(-90, 90), fit_reg=False, marginal_kws={'color': '#ff0000'}
)
pyplot.setp(graph_1.ax_marg_y.patches, color='#0000ff')
pyplot.rc('legend')
pyplot.xlabel('Longitude')
pyplot.ylabel('Latitude')
pyplot.title('Craters Location Full Dataset')
pyplot.tick_params(axis='both')
pyplot.show()
graph_1.savefig('output_location_full_dataset.png')

# creating new subset
working_subset = data[
    (data['DEPTH_RIMFLOOR_TOPOG'] > 0) & (data['DEPTH_RIMFLOOR_TOPOG'] <= 3)
    & (data['DIAM_CIRCLE_IMAGE'] > 0) & (data['DIAM_CIRCLE_IMAGE'] <= 100)
]

graph_2 = seaborn.jointplot(
    x='LONGITUDE_CIRCLE_IMAGE', y='LATITUDE_CIRCLE_IMAGE', data=working_subset, kind='reg', color='#774499', dropna=True,
    height=13, space=0.3, ratio=4, xlim=(-180, 180), ylim=(-90, 90), fit_reg=False, marginal_kws={'color': '#ff0000'}
)
pyplot.setp(graph_2.ax_marg_y.patches, color='#0000ff')
pyplot.rc('legend')
pyplot.xlabel('Longitude')
pyplot.ylabel('Latitude')
pyplot.title('Craters Location Working Dataset')
pyplot.tick_params(axis='both')
pyplot.show()
graph_2.savefig('output_location_working_dataset.png')


# frequency distribution for variable DEPTH_GROUP
depth = [0, 0.3, 0.6, 0.9, 1.2, 1.5, 1.8, 2.1, 2.4, 2.7, 3.0]
depth_labels = ['0 - 0.3', '0.3 - 0.6', '0.6 - 0.9', '0.9 - 1.2', '1.2 - 1.5', '1.5 - 1.8', '1.8 - 2.1', '2.1 - 2.4', '2.4 - 2.7', '2.7 - 3.0']
working_subset['DEPTH_GROUP'] = pandas.cut(working_subset.DEPTH_RIMFLOOR_TOPOG, bins=depth, labels=depth_labels, right=True)

print()
print('=========================================')
print('Value counts for varaible - DEPTH_GROUP: ')
print('=========================================')
count_depth_group = working_subset['DEPTH_GROUP'].value_counts(sort=False)
print(count_depth_group)
print()

# Percentages for DEPTH_GROUP Variable
print()
print('========================================')
print('Percentages for varaible - DEPTH_GROUP: ')
print('========================================')
percentages_depth_group = working_subset['DEPTH_GROUP'].value_counts(sort=False, normalize=True)
print(percentages_depth_group)
print()

print()
print('=========')
print('Results: ')
print('=========')
data_frame_1 = pandas.DataFrame(numpy.array(count_depth_group), index=count_depth_group.index.values, columns=['Frequency'])
data_frame_2 = pandas.DataFrame(numpy.array(percentages_depth_group * 100), index=percentages_depth_group.index.values, columns=['Percentage'])
output_depth_group = pandas.concat([data_frame_1, data_frame_2], axis=1)

output_depth_group['Frequency'] = output_depth_group.Frequency
output_depth_group['Percentage'] = output_depth_group.Percentage

output_depth_group.index.name = 'DEPTH_GROUP'
print(output_depth_group)
print()



# frequency distribution for variable DIAMETER_GROUP
diameter = [i for i in range(0, 105, 5)]
diameter_labels = ['{0} - {1}'.format(i, i+5) for i in range(0, 100, 5)]
working_subset['DIAMETER_GROUP'] = pandas.cut(
    working_subset.DIAM_CIRCLE_IMAGE,
    bins=diameter,
    labels=diameter_labels,
    right=True
)

print()
print('============================================')
print('Value counts for varaible - DIAMETER_GROUP: ')
print('============================================')
count_diameter_group = working_subset['DIAMETER_GROUP'].value_counts(sort=False)
print(count_diameter_group)
print()

# Percentages for DIAMETER_GROUP Variable
print()
print('===========================================')
print('Percentages for varaible - DIAMETER_GROUP: ')
print('===========================================')
percentages_diameter_group = working_subset['DIAMETER_GROUP'].value_counts(sort=False, normalize=True)
print(percentages_diameter_group)
print()

print()
print('=========')
print('Results: ')
print('=========')
data_frame_3 = pandas.DataFrame(numpy.array(count_diameter_group), index=count_diameter_group.index.values, columns=['Frequency'])
data_frame_4 = pandas.DataFrame(numpy.array(percentages_diameter_group * 100), index=percentages_diameter_group.index.values, columns=['Percentage'])
output_diameter_group = pandas.concat([data_frame_3, data_frame_4], axis=1)

output_diameter_group['Frequency'] = output_diameter_group.Frequency
output_diameter_group['Percentage'] = output_diameter_group.Percentage

output_diameter_group.index.name = 'DIAMETER_GROUP'
print(output_diameter_group)
print()



# frequency distribution for variable LATITUDE_GROUP
latitude = [i for i in range(-90, 105, 15)]
latitude_labels = ['{0} - {1}'.format(i, i+15) for i in range(-90, 90, 15)]
working_subset['LATITUDE_GROUP'] = pandas.cut(working_subset.LATITUDE_CIRCLE_IMAGE, bins=latitude,
                                              labels=latitude_labels, right=True)

print()
print('============================================')
print('Value counts for varaible - LATITUDE_GROUP: ')
print('============================================')
count_latitude_group = working_subset['LATITUDE_GROUP'].value_counts(sort=False)
print(count_latitude_group)
print()

# Percentages for LATITUDE_GROUP Variable
print()
print('===========================================')
print('Percentages for varaible - LATITUDE_GROUP: ')
print('===========================================')
percentages_latitude_group = working_subset['LATITUDE_GROUP'].value_counts(sort=False, normalize=True)
print(percentages_latitude_group)
print()

print()
print('=========')
print('Results: ')
print('=========')
data_frame_5 = pandas.DataFrame(numpy.array(count_latitude_group), index=count_latitude_group.index.values, columns=['Frequency'])
data_frame_6 = pandas.DataFrame(numpy.array(percentages_latitude_group * 100), index=percentages_latitude_group.index.values, columns=['Percentage'])
output_latitude_group = pandas.concat([data_frame_5, data_frame_6], axis=1)

output_latitude_group['Frequency'] = output_latitude_group.Frequency
output_latitude_group['Percentage'] = output_latitude_group.Percentage

output_latitude_group.index.name = 'LATITUDE_GROUP'
print(output_latitude_group)
print()


# creating an workbook
excel_writer = pandas.ExcelWriter('Mars_Craters_WorkBook_Assignment_3.xlsx')
output_depth_group.to_excel(excel_writer, sheet_name='DEPTH_GROUP', float_format='%0.4f', startrow=1, startcol=0)
output_diameter_group.to_excel(excel_writer, sheet_name='DIAMETER_GROUP', float_format='%0.4f', startrow=1, startcol=0)
output_latitude_group.to_excel(excel_writer, sheet_name='LATITUDE_GROUP', float_format='%0.4f', startrow=1, startcol=0)
excel_writer.save()

wb = openpyxl.load_workbook('Mars_Craters_WorkBook_Assignment_3.xlsx')
sheets = wb.sheetnames

for i in range(len(sheets)):
    ws = wb.worksheets[i]
    variable_name = ws.cell(row=2, column=1).value
    title = 'Mars Crater Dataset Frequency Distribution for variable: ' + variable_name
    ws.cell(row=1, column=1).value = title.upper()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    wb.save('Mars_Craters_WorkBook_Assignment_3.xlsx')
